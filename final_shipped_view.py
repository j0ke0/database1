import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psycopg2
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from image_icon_paths import icon_path, dbnames, users, passwords, hosts, ports, image_path

def get_database_cursor():
    try:
        db_password = 'Nolcubs0#'
        db_connection = psycopg2.connect(
            dbname=dbnames,
            user=users,
            password=passwords,
            host=hosts,
            port=ports
        )
        return db_connection.cursor()
    except psycopg2.Error as e:
        messagebox.showerror("Database Connection Error", f"Failed to connect to the database: {e}")

def populate_table(tree, db_cursor):
    try:
        # Clear old data from the Treeview
        tree.delete(*tree.get_children())

        # Execute the SQL query to fetch data from the shipped_final table
        db_cursor.execute('SELECT customer_name, shipping_number, "logged_in_users", time_stamp, serials FROM public.shipped_final;')
        results = db_cursor.fetchall()

        # Insert data into the table
        for row in results:
            # Format the timestamp to exclude milliseconds
            formatted_row = list(row)
            formatted_row[-2] = row[-2].strftime('%d-%m-%Y - %H:%M:%S')
            tree.insert("", "end", values=formatted_row)

    except psycopg2.Error as e:
        messagebox.showerror("Database Error", f"Error executing the query: {e}")

def export_to_excel(db_cursor):
    try:
        # Execute the SQL query to fetch data from the shipped_final table
        db_cursor.execute('SELECT customer_name, shipping_number, "logged_in_users", time_stamp, serials FROM public.shipped_final;')
        results = db_cursor.fetchall()

        # Get current date without time
        current_date = datetime.now().strftime('%Y-%m-%d')

        # Extract customer_name, shipping_number, and date (without time) from the first row of results
        if results:
            customer_name, shipping_number, _, date_with_time, _ = results[0]
            date_without_time = date_with_time.strftime('%d-%m-%Y')
        else:
            # If there are no results, use default values
            customer_name, shipping_number, date_without_time = "Unknown", "Unknown", current_date

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active

        # Bold font for headers
        bold_font = Font(bold=True)
        # Center alignment for all cells
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Write column headers and apply bold font and center alignment
        headers = ["Customer Name", "Shipping Number", "Logged-in User", "Date & Time", "Serials"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = bold_font
            ws.cell(row=1, column=col_num).alignment = centered_alignment

        # Write data rows and apply center alignment
        for row_idx, row in enumerate(results, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).alignment = centered_alignment

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Freeze the top row (headers)
        ws.freeze_panes = 'A2'

        # Construct the file name
        file_name = f"{customer_name}_{shipping_number}_{date_without_time}.xlsx"

        # Ask user for file save location
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=file_name)
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Export Successful", "Data exported to Excel successfully!")
        else:
            messagebox.showinfo("Export Cancelled", "Export to Excel was cancelled.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred during export: {e}")


def search_tree(tree, input_text):
    # Retrieve the search query from the input text widget
    query = input_text.get().strip().lower()

    # If the query is empty, reset the Treeview to display all rows
    if not query:
        tree.selection_remove(tree.get_children())
        tree.yview_moveto(0)  # Scroll back to the top
        return

    # Store the indices of matching items
    matching_items = []

    # Iterate through all items in the Treeview
    for item in tree.get_children():
        # Retrieve values of each item
        values = tree.item(item, 'values')

        # Check if any value in the item matches the search query
        if any(query in str(value).lower() for value in values):
            matching_items.append(item)
            tree.selection_add(item)
        else:
            tree.selection_remove(item)

    # Scroll to the middle of the first matching item (if any)
    if matching_items:
        first_matching_item = matching_items[0]
        tree.yview_moveto(tree.index(first_matching_item) / len(tree.get_children()))

def final_view_shipped(root):
    try:
        # Create a new Toplevel window
        data_popup = tk.Toplevel(root)
        data_popup.title("Shipped History")

        data_popup.iconbitmap(icon_path)


        # Create a Treeview widget
        tree = ttk.Treeview(data_popup, columns=("Customer Name", "Shipping Number", "Logged-in User", "Date & Time", "Serials"), show="headings")
        tree.grid(row=0, column=0, sticky="nsew")

        # Set up column headings with centered text
        for col in ("Customer Name", "Shipping Number", "Logged-in User", "Date & Time", "Serials"):
            tree.heading(col, text=col, anchor="center")

        # Center align the data inside the tree
        for col in ("Customer Name", "Shipping Number", "Logged-in User", "Date & Time", "Serials"):
            tree.column(col, anchor="center")

        # Adjust the width of the "Serials" column
        tree.column("Serials", width=600)  # Adjust the width as needed

        # Establish database connection and cursor
        db_cursor = get_database_cursor()

        # Populate table when window is opened
        data_popup.bind("<Map>", lambda event: populate_table(tree, db_cursor))

        # Close database cursor and connection when window is destroyed
        data_popup.bind("<Destroy>", lambda event: db_cursor.close() if db_cursor else None)
        data_popup.bind("<Destroy>", lambda event: db_cursor.connection.close() if db_cursor else None)

        # Create a frame to hold the buttons and input text
        button_frame = tk.Frame(data_popup)
        button_frame.grid(row=1, column=0, pady=5)

        # Add export to Excel button
        export_button = ttk.Button(button_frame, text="Export to Excel", command=lambda: export_to_excel(db_cursor))
        export_button.grid(row=0, column=0, padx=5)

        # Add a cancel button
        cancel_button = ttk.Button(button_frame, text="Cancel", command=data_popup.destroy)
        cancel_button.grid(row=0, column=3, padx=5)

        # Add another button (example button)
        another_button = ttk.Button(button_frame, text="Find Serial", command=lambda: search_tree(tree, input_text))
        another_button.grid(row=0, column=1, padx=5)

        # Add an input text widget
        input_text = ttk.Entry(button_frame)
        input_text.grid(row=0, column=2, padx=5)

        # Configure grid weights to make treeview expandable
        data_popup.grid_rowconfigure(0, weight=1)
        data_popup.grid_columnconfigure(0, weight=1)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")



