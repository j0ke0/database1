import re
import os
import openpyxl
import psycopg2
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import filedialog
from datetime import datetime
from openpyxl.styles import Alignment, Font
import tkinter.messagebox as messagebox
from tkinter import messagebox, OptionMenu
from image_icon_paths import icon_path, dbnames, users, passwords, hosts, ports, image_path

db_password = 'Nolcubs0#'
db_connection = psycopg2.connect(
    dbname=dbnames,
    user=users,
    password=passwords,
    host=hosts,
    port=ports
)
db_cursor = db_connection.cursor()

def submit_and_display_data(entry_fields, logged_in_user, popup):
    db_cursor = None
    try:
        db_connection = psycopg2.connect(
            dbname=dbnames,
            user=users,
            password=passwords,
            host=hosts,
            port=ports
        )
        db_cursor = db_connection.cursor()

        data = []
        for entry in entry_fields:
            if isinstance(entry, tk.Text):
                value = entry.get("1.0", "end-1c")
            else:
                value = entry.get()
            value = value.strip().upper()
            if len(value) <= 5:
                popup.grab_set()
                raise ValueError("All text must be longer than 6 characters.")
            data.append(value)

        customer_name, shipping_number, serials = data

        logged_in_users = logged_in_user.split(":")[1].strip()

        if not all(data):
            popup.grab_set()
            raise ValueError("All fields must be filled out.")

        query_pagadmin4 = '''
            INSERT INTO shipped_final (customer_Name, shipping_number, logged_in_users, time_stamp, serials)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP,  %s)
        '''
        db_cursor.execute(query_pagadmin4, (customer_name, shipping_number, logged_in_users, serials))

        db_connection.commit()
        popup.grab_set()  # Grab the focus to the popup window
        messagebox.showinfo("Success", "Data entry successfully")
    except psycopg2.Error as e:
        if db_connection:
            db_connection.rollback()
        messagebox.showerror("Database Error", f"Error inserting data: {e}")
    except ValueError as ve:
        messagebox.showerror("Input Error", str(ve))
    finally:
        if db_cursor:
            db_cursor.close()
        if db_connection:
            db_connection.close()

def disable_enter(event):
    event.widget.insert("insert", " ")  # Insert a space character
    return "break"

def final_shipped_goods(root, logged_in_user):
    popup = tk.Toplevel(root)
    popup.title("Shipping Goods")

    popup.iconbitmap(icon_path)

    popup.iconbitmap(icon_path)

    add_path = "shipping.png"
    full_path = image_path + add_path

    image = Image.open(full_path)
    photo = ImageTk.PhotoImage(image)
    image_label = tk.Label(popup, image=photo)
    image_label.grid(row=0, column=0, columnspan=2, pady=8)

    entry_labels = ["Customer Name:", "Shipping Number:", "Serials:"]
    entry_fields = []

    bold_font = ('Arial', 10, 'bold')  # Define bold font style

    for i, label_text in enumerate(entry_labels):
        label = tk.Label(popup, text=label_text, font=bold_font)  # Set bold font for label
        label.grid(row=i+1, column=0, sticky='e')
        if label_text == "Serials:":
            entry = tk.Text(popup, height=3, width=19, wrap=tk.WORD)  # Enable word wrapping
            entry.grid(row=i+1, column=1, sticky='w')  # Grid the Text widget
            # Bind the Return key to disable_enter function
            entry.bind("<Return>", disable_enter)
        else:
            entry = tk.Entry(popup)
            entry.grid(row=i+1, column=1, sticky='w')
        entry_fields.append(entry)

    submit_button = ttk.Button(popup, text="Submit", command=lambda: submit_and_display_data(entry_fields, logged_in_user, popup))
    submit_button.grid(row=len(entry_labels) + 2, column=0, padx=7, pady=8, sticky="ew")


    def get_last_row_from_database():
        try:
            db_cursor.execute("SELECT * FROM shipped_final ORDER BY time_stamp DESC LIMIT 1")
            last_row = db_cursor.fetchone()
            return last_row
        except psycopg2.Error as e:
            messagebox.showerror("Database Error", f"Error fetching data from database: {e}")
            return None

    def format_timestamp(timestamp):
        return timestamp.strftime("%d %b %Y - %H:%M:%S")  # Format as "day month year time"

    def create_excel_file_from_database():
        last_row = get_last_row_from_database()
        if last_row:
            customer_name = last_row[1]  # Assuming customer name is the second column after the primary key
            shipping_number = last_row[2]  # Assuming shipping number is the third column after the primary key

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Shipped Data"

            headers = ["Table Number", "Customer Name", "Shipping Number", "Logged-in User", "Date & Time", "Serials"]
            bold_font = Font(bold=True)  # Create a bold font object

            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = bold_font  # Apply bold font to header cells

            last_row_formatted = list(last_row)
            last_row_formatted[-2] = format_timestamp(last_row[-2])  # Format timestamp
            sheet.append(last_row_formatted)

            # Center align all cells
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths based on content
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                sheet.column_dimensions[column].width = adjusted_width

            file_name = f"{customer_name}_{shipping_number}.xlsx"
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=file_name)
            if file_path:
                workbook.save(file_path)
                messagebox.showinfo("Excel Created", f"Excel file '{file_path}' created successfully.")
                popup.destroy()  # Destroy the popup window after successful Excel creation
            else:
                messagebox.showinfo("File Not Saved", "No file selected. Excel file not created.")
        else:
            messagebox.showerror("No Data", "No data available to export to Excel.")



    create_excel = ttk.Button(popup, text="Create Excel", width=12, command=create_excel_file_from_database)
    create_excel.grid(row=len(entry_labels) + 2, column=1, padx=7, pady=8)


    popup.mainloop()
